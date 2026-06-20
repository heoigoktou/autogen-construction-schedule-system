"""Read real-case source documents and expose auditable evidence rows."""

from __future__ import annotations

import csv
import re
import shutil
import subprocess
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook

from tools.dwg_tools import convert_dwg_to_dxf, ensure_oda_converter, read_dxf_text

SUPPORTED_DOCUMENT_EXTENSIONS = {
    ".txt",
    ".md",
    ".csv",
    ".xlsx",
    ".xlsm",
    ".doc",
    ".docx",
    ".pdf",
    ".dwg",
    ".dxf",
}

CN_TZ = timezone(timedelta(hours=8))
MOJIBAKE_MARKERS = (
    "\ufffd",
    "\u951f",  # Chinese mojibake marker often rendered as "kun".
    "\u62f7",
    "\u951b",
    "\u6b7f",
    "\u93c2",
    "\u7481",
    "\u9435",
)


@dataclass(frozen=True)
class SourceDocument:
    """Plain-text plus optional structured evidence extracted from one file."""

    path: Path
    text: str
    warning: str = ""
    sections: tuple[dict[str, str], ...] = ()
    tables: tuple[dict[str, str], ...] = ()

    @property
    def name(self) -> str:
        return self.path.name


def read_source_documents(
    input_dir: str | Path,
    *,
    dwg_conversion_dir: str | Path | None = None,
    install_oda_if_missing: bool = False,
    dwg_timeout_seconds: int = 120,
) -> list[SourceDocument]:
    """Read supported files under an input directory."""

    root = Path(input_dir)
    if not root.exists():
        return []

    return [
        read_source_document(
            path,
            dwg_conversion_dir=dwg_conversion_dir,
            install_oda_if_missing=install_oda_if_missing,
            dwg_timeout_seconds=dwg_timeout_seconds,
        )
        for path in sorted(_iter_supported_files(root))
    ]


def read_source_document(
    path: str | Path,
    *,
    dwg_conversion_dir: str | Path | None = None,
    install_oda_if_missing: bool = False,
    dwg_timeout_seconds: int = 120,
) -> SourceDocument:
    """Read one supported source document."""

    file_path = Path(path)
    suffix = file_path.suffix.lower()
    try:
        if suffix in {".txt", ".md"}:
            return _document_from_text(file_path, _read_text_file(file_path))
        if suffix == ".csv":
            return _document_from_text(file_path, _read_csv(file_path))
        if suffix in {".xlsx", ".xlsm"}:
            return _document_from_text(file_path, _read_xlsx(file_path))
        if suffix == ".doc":
            return _read_doc(file_path)
        if suffix == ".docx":
            text, sections, tables = _read_docx_parts(file_path)
            return SourceDocument(file_path, text, sections=sections, tables=tables)
        if suffix == ".pdf":
            return _document_from_text(file_path, _read_pdf(file_path))
        if suffix == ".dxf":
            return _document_from_text(file_path, read_dxf_text(file_path))
        if suffix == ".dwg":
            return _read_dwg(
                file_path,
                dwg_conversion_dir=dwg_conversion_dir,
                install_oda_if_missing=install_oda_if_missing,
                dwg_timeout_seconds=dwg_timeout_seconds,
            )
    except Exception as exc:  # pragma: no cover - defensive branch for bad source files
        return SourceDocument(file_path, "", f"read failed: {exc}")
    return SourceDocument(file_path, "", f"unsupported file type: {suffix}")


def concatenate_documents(
    documents: Iterable[SourceDocument],
    *,
    max_chars: int = 30000,
) -> str:
    """Concatenate documents with source names and a conservative length cap."""

    chunks: list[str] = []
    used = 0
    for document in documents:
        body = document.warning if document.warning and not document.text else document.text.strip()
        section = f"[file: {document.name}]\n{body}\n"
        if used + len(section) > max_chars:
            remaining = max_chars - used
            if remaining > 200:
                chunks.append(section[:remaining] + "\n[content truncated]\n")
            break
        chunks.append(section)
        used += len(section)
    return "\n".join(chunks).strip()


def build_document_evidence_rows(
    documents: Iterable[SourceDocument],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """Build rows for document_sections and document_tables."""

    created_at = _now_iso()
    section_rows: list[dict[str, str]] = []
    table_rows: list[dict[str, str]] = []
    for doc_index, document in enumerate(documents, start=1):
        sections = document.sections or _sections_from_text(document.text)
        for index, section in enumerate(sections, start=1):
            normalized = _normalize_evidence_text(
                section.get("normalized_text") or section.get("raw_text") or ""
            )
            if not normalized:
                continue
            section_rows.append(
                {
                    "evidence_id": section.get("evidence_id") or f"SEC-{doc_index:03d}-{index:04d}",
                    "document_name": document.name,
                    "source_type": section.get("source_type") or "section",
                    "section_title": section.get("section_title") or _guess_section_title(normalized),
                    "page_or_order": str(section.get("page_or_order") or index),
                    "raw_text": _truncate(section.get("raw_text") or normalized, 1800),
                    "normalized_text": _truncate(normalized, 1800),
                    "created_by": "document_parser",
                    "created_at": created_at,
                }
            )
        for index, table in enumerate(document.tables, start=1):
            normalized = _normalize_evidence_text(
                table.get("normalized_text") or table.get("raw_text") or ""
            )
            if not normalized:
                continue
            table_rows.append(
                {
                    "evidence_id": table.get("evidence_id") or f"TAB-{doc_index:03d}-{index:04d}",
                    "document_name": document.name,
                    "source_type": table.get("source_type") or "table",
                    "section_title": table.get("section_title") or "table",
                    "page_or_order": str(table.get("page_or_order") or index),
                    "raw_text": _truncate(table.get("raw_text") or normalized, 1800),
                    "normalized_text": _truncate(normalized, 1800),
                    "created_by": "document_parser",
                    "created_at": created_at,
                }
            )
    return section_rows, table_rows


def has_mojibake(text: str) -> bool:
    """Return True when text looks like decoded Chinese mojibake."""

    if not text:
        return False
    sample = text[:6000]
    marker_hits = sum(sample.count(marker) for marker in MOJIBAKE_MARKERS)
    chinese_chars = sum(1 for char in sample if "\u4e00" <= char <= "\u9fff")
    ascii_letters = sum(1 for char in sample if char.isascii() and char.isalpha())
    if marker_hits >= 8:
        return True
    return marker_hits >= 3 and chinese_chars < max(marker_hits * 12, ascii_letters // 4, 60)


def _document_from_text(path: Path, text: str, warning: str = "") -> SourceDocument:
    return SourceDocument(path, text, warning=warning, sections=_sections_from_text(text))


def _iter_supported_files(root: Path) -> Iterable[Path]:
    for path in root.rglob("*"):
        if path.is_file() and path.suffix.lower() in SUPPORTED_DOCUMENT_EXTENSIONS:
            if any(part in {".doc_conversion", ".dwg_conversion"} for part in path.parts):
                continue
            if path.name.startswith("~$") or path.name.lower() == "readme.md":
                continue
            yield path


def _read_text_file(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(errors="ignore")


def _read_csv(path: Path) -> str:
    content = _read_text_file(path)
    rows = []
    for row in csv.reader(content.splitlines()):
        rows.append(" | ".join(cell.strip() for cell in row if str(cell).strip()))
    return "\n".join(rows)


def _read_xlsx(path: Path) -> str:
    workbook = load_workbook(path, data_only=True, read_only=True)
    lines: list[str] = []
    try:
        for sheet in workbook.worksheets:
            lines.append(f"## sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [
                    str(value).strip()
                    for value in row
                    if value is not None and str(value).strip()
                ]
                if values:
                    lines.append(" | ".join(values))
    finally:
        workbook.close()
    return "\n".join(lines)


def _read_docx(path: Path) -> str:
    text, _, _ = _read_docx_parts(path)
    return text


def _read_docx_parts(path: Path) -> tuple[str, tuple[dict[str, str], ...], tuple[dict[str, str], ...]]:
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(path) as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml_bytes)
    body = root.find("w:body", namespace)
    if body is None:
        return "", (), ()

    paragraphs: list[str] = []
    section_rows: list[dict[str, str]] = []
    table_rows: list[dict[str, str]] = []
    paragraph_order = 0
    table_order = 0
    for child in list(body):
        tag = _local_name(child.tag)
        if tag == "p":
            text = _node_text(child, namespace)
            if not text:
                continue
            paragraph_order += 1
            paragraphs.append(text)
            section_rows.append(
                {
                    "source_type": "paragraph",
                    "section_title": _guess_section_title(text),
                    "page_or_order": str(paragraph_order),
                    "raw_text": text,
                    "normalized_text": _normalize_evidence_text(text),
                }
            )
        elif tag == "tbl":
            lines = []
            for table_row in child.findall(".//w:tr", namespace):
                cells = [
                    _normalize_evidence_text(_node_text(cell, namespace))
                    for cell in table_row.findall("./w:tc", namespace)
                ]
                cells = [cell for cell in cells if cell]
                if cells:
                    lines.append(" | ".join(cells))
            if lines:
                table_order += 1
                table_text = "\n".join(lines)
                paragraphs.append(table_text)
                table_rows.append(
                    {
                        "source_type": "table",
                        "section_title": "table",
                        "page_or_order": str(table_order),
                        "raw_text": table_text,
                        "normalized_text": _normalize_evidence_text(table_text),
                    }
                )
    return "\n".join(paragraphs), tuple(section_rows), tuple(table_rows)


def _read_doc(path: Path) -> SourceDocument:
    """Read legacy Word .doc files, preferring conversion to OOXML."""

    warnings: list[str] = []
    fallback: SourceDocument | None = None
    for converter in (_convert_doc_with_word_com, _convert_doc_with_libreoffice):
        converted, warning = converter(path)
        if warning:
            warnings.append(warning)
        if not converted:
            continue
        try:
            text, sections, tables = _read_docx_parts(converted)
        except Exception as exc:  # pragma: no cover - depends on local converter output
            warnings.append(f"converted docx parse failed: {exc}")
            continue
        document = SourceDocument(
            path,
            text,
            warning="; ".join(warnings),
            sections=sections,
            tables=tables,
        )
        if text and not has_mojibake(text):
            return document
        if text and fallback is None:
            fallback = document

    text, warning = _read_doc_with_word_com_text(path)
    if warning:
        warnings.append(warning)
    if text:
        document = _document_from_text(path, text, warning="; ".join(warnings))
        if not has_mojibake(text):
            return document
        fallback = fallback or document

    if fallback is not None:
        joined = "; ".join(item for item in (fallback.warning, "possible mojibake") if item)
        return SourceDocument(
            path,
            fallback.text,
            joined,
            sections=fallback.sections,
            tables=fallback.tables,
        )
    return SourceDocument(
        path,
        "",
        "; ".join(warnings) or "legacy .doc read failed: no Word or LibreOffice converter",
    )


def _convert_doc_with_word_com(path: Path) -> tuple[Path | None, str]:
    """Convert .doc to .docx with Word COM when available."""

    try:
        import win32com.client  # type: ignore[import-not-found]
    except ModuleNotFoundError:
        return None, "pywin32 is not installed; Word COM conversion skipped"

    output_dir = path.parent / ".doc_conversion"
    output_dir.mkdir(parents=True, exist_ok=True)
    converted = output_dir / f"{path.stem}.docx"
    word = None
    document = None
    try:
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        document = word.Documents.Open(str(path.resolve()), ReadOnly=True)
        document.SaveAs2(str(converted.resolve()), FileFormat=16)
        return converted if converted.exists() else None, f"converted by Word COM: {converted.name}"
    except Exception as exc:  # pragma: no cover - depends on local Word installation
        return None, f"Word COM conversion failed: {exc}"
    finally:
        if document is not None:
            try:
                document.Close(False)
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass


def _convert_doc_with_libreoffice(path: Path) -> tuple[Path | None, str]:
    """Convert .doc to .docx with LibreOffice when available."""

    executable = _find_libreoffice()
    if not executable:
        return None, "LibreOffice/soffice not found; conversion skipped"

    output_dir = path.parent / ".doc_conversion"
    output_dir.mkdir(parents=True, exist_ok=True)
    command = [
        executable,
        "--headless",
        "--convert-to",
        "docx",
        "--outdir",
        str(output_dir),
        str(path.resolve()),
    ]
    try:
        subprocess.run(command, check=True, capture_output=True, text=True, timeout=120)
        converted = output_dir / f"{path.stem}.docx"
        if not converted.exists():
            return None, "LibreOffice conversion did not produce a docx file"
        return converted, f"converted by LibreOffice: {converted.name}"
    except Exception as exc:  # pragma: no cover - depends on local LibreOffice installation
        return None, f"LibreOffice conversion failed: {exc}"


def _read_doc_with_word_com_text(path: Path) -> tuple[str, str]:
    """Fallback direct text extraction from .doc with Word COM."""

    try:
        import win32com.client  # type: ignore[import-not-found]
    except ModuleNotFoundError:
        return "", "pywin32 is not installed; direct Word COM read skipped"

    word = None
    document = None
    try:
        word = win32com.client.Dispatch("Word.Application")
        word.Visible = False
        document = word.Documents.Open(str(path.resolve()), ReadOnly=True)
        return str(document.Content.Text or "").strip(), "read by direct Word COM text fallback"
    except Exception as exc:  # pragma: no cover - depends on local Word installation
        return "", f"direct Word COM read failed: {exc}"
    finally:
        if document is not None:
            try:
                document.Close(False)
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass


def _find_libreoffice() -> str:
    for name in ("soffice", "libreoffice"):
        executable = shutil.which(name)
        if executable:
            return executable
    common = [
        Path("C:/Program Files/LibreOffice/program/soffice.exe"),
        Path("C:/Program Files (x86)/LibreOffice/program/soffice.exe"),
    ]
    for path in common:
        if path.exists():
            return str(path)
    return ""


def _read_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ModuleNotFoundError:
        return "PDF read requires pypdf: python -m pip install pypdf"

    reader = PdfReader(str(path))
    pages = []
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"## page {index}\n{text.strip()}")
    return "\n\n".join(pages)


def _read_dwg(
    path: Path,
    *,
    dwg_conversion_dir: str | Path | None,
    install_oda_if_missing: bool,
    dwg_timeout_seconds: int,
) -> SourceDocument:
    converter, warning = ensure_oda_converter(
        install_if_missing=install_oda_if_missing,
        timeout_seconds=max(dwg_timeout_seconds, 120),
    )
    if converter is None:
        return SourceDocument(path, "", warning)

    output_dir = Path(dwg_conversion_dir) if dwg_conversion_dir else path.parent / ".dwg_conversion"
    try:
        dxf_path = convert_dwg_to_dxf(
            path,
            output_dir=output_dir,
            converter_path=converter,
            timeout_seconds=dwg_timeout_seconds,
        )
        text = read_dxf_text(dxf_path)
        return _document_from_text(path, f"converted from DWG: {dxf_path.name}\n\n{text}")
    except Exception as exc:
        return SourceDocument(path, "", f"DWG to DXF conversion or parse failed: {exc}")


def _node_text(node: ElementTree.Element, namespace: dict[str, str]) -> str:
    parts = [
        text_node.text or ""
        for text_node in node.findall(".//w:t", namespace)
        if text_node.text and text_node.text.strip()
    ]
    return "".join(parts).strip()


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _sections_from_text(text: str) -> tuple[dict[str, str], ...]:
    paragraphs = [item.strip() for item in re.split(r"[\r\n]+", text or "") if item.strip()]
    if not paragraphs and text.strip():
        paragraphs = [text.strip()]

    sections: list[dict[str, str]] = []
    buffer: list[str] = []
    title = ""
    order = 0

    def flush() -> None:
        nonlocal order, buffer, title
        if not buffer:
            return
        order += 1
        raw = "\n".join(buffer)
        sections.append(
            {
                "source_type": "section",
                "section_title": title or _guess_section_title(raw),
                "page_or_order": str(order),
                "raw_text": raw,
                "normalized_text": _normalize_evidence_text(raw),
            }
        )
        buffer = []

    for paragraph in paragraphs:
        if _looks_like_heading(paragraph) and buffer:
            flush()
            title = paragraph[:80]
        buffer.append(paragraph)
        if len("\n".join(buffer)) > 1200:
            flush()
    flush()
    return tuple(sections)


def _looks_like_heading(text: str) -> bool:
    compact = text.strip()
    if not compact or len(compact) > 80:
        return False
    return bool(
        re.match(r"^\u7b2c[\u4e00-\u9fff0-9]+[\u7ae0\u8282\u6761]", compact)
        or re.match(r"^\d+(?:\.\d+){0,3}\s*\S+", compact)
        or compact.endswith(
            (
                "\u5de5\u7a0b\u6982\u51b5",
                "\u65bd\u5de5\u90e8\u7f72",
                "\u65bd\u5de5\u8fdb\u5ea6\u8ba1\u5212",
                "\u8d44\u6e90\u8ba1\u5212",
                "\u65bd\u5de5\u65b9\u6848",
            )
        )
    )


def _guess_section_title(text: str) -> str:
    stripped = (text or "").strip()
    first_line = stripped.splitlines()[0] if stripped else ""
    return first_line[:80] if _looks_like_heading(first_line) else ""


def _normalize_evidence_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").replace("\u3000", " ")).strip()


def _truncate(text: str, max_chars: int) -> str:
    value = str(text or "")
    return value if len(value) <= max_chars else value[: max_chars - 12] + "...<truncated>"


def _now_iso() -> str:
    return datetime.now(CN_TZ).isoformat(timespec="seconds")
