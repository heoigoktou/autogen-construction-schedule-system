"""Visualization helpers for generated construction schedules."""

from __future__ import annotations

import json
import math
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import networkx as nx
from openpyxl import load_workbook
from matplotlib import font_manager

from blackboard.excel_store import ExcelBlackboardStore
from tools.resource_tools import build_resource_load
from tools.schedule_tools import build_initial_schedule


DEFAULT_OUTPUT_DIRNAME = "visualizations"
CRITICAL_COLOR = "#D94841"
NORMAL_COLOR = "#4E79A7"
FLOAT_COLOR = "#59A14F"
CAPACITY_COLOR = "#F28E2B"
GRID_COLOR = "#D9DEE7"
TEXT_COLOR = "#263238"


@dataclass(frozen=True)
class VisualizationResult:
    """Summary of generated visualization artifacts."""

    output_dir: Path
    artifacts: dict[str, Path]
    warnings: list[str]

    def to_dict(self) -> dict[str, Any]:
        return {
            "output_dir": str(self.output_dir),
            "artifacts": {name: str(path) for name, path in self.artifacts.items()},
            "warnings": self.warnings,
        }


def generate_schedule_visualizations(
    store: ExcelBlackboardStore,
    output_dir: str | Path,
    *,
    title: str = "AgentChat Schedule Visualization",
) -> VisualizationResult:
    """Generate Gantt, CPM, float, and resource load visuals from blackboard tables."""

    return generate_schedule_visualizations_from_rows(
        {
            "schedule_initial": store.read_rows("schedule_initial"),
            "cpm_analysis": store.read_rows("cpm_analysis"),
            "network_edges": store.read_rows("network_edges"),
            "resource_load_daily": store.read_rows("resource_load_daily"),
        },
        output_dir,
        title=title,
    )


def generate_schedule_visualizations_from_export_dir(
    schedule_dir: str | Path,
    output_dir: str | Path,
    *,
    title: str = "AgentChat Schedule Visualization",
) -> VisualizationResult:
    """Generate schedule visuals from standalone exported Excel files."""

    rows = read_exported_schedule_tables(schedule_dir)
    return generate_schedule_visualizations_from_rows(rows, output_dir, title=title)


def generate_schedule_visualizations_from_rows(
    rows: dict[str, list[dict[str, Any]]],
    output_dir: str | Path,
    *,
    title: str = "AgentChat Schedule Visualization",
) -> VisualizationResult:
    """Generate schedule visuals from already-loaded table rows."""

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    _configure_matplotlib()

    schedule_rows = rows.get("schedule_initial") or []
    cpm_rows = rows.get("cpm_analysis") or []
    edge_rows = rows.get("network_edges") or []
    resource_load_rows = rows.get("resource_load_daily") or []

    artifacts: dict[str, Path] = {}
    warnings: list[str] = []

    if schedule_rows:
        artifacts["gantt_chart"] = write_gantt_chart(
            output_path / "gantt_chart.png",
            schedule_rows,
            title=f"{title} - Gantt",
        )
    else:
        warnings.append("schedule_initial is empty; Gantt chart was skipped.")

    if schedule_rows and cpm_rows:
        artifacts["cpm_float_chart"] = write_cpm_float_chart(
            output_path / "cpm_float_chart.png",
            schedule_rows,
            cpm_rows,
            title=f"{title} - CPM Float",
        )
        artifacts["cpm_network"] = write_cpm_network_chart(
            output_path / "cpm_network.png",
            schedule_rows,
            cpm_rows,
            edge_rows,
            title=f"{title} - CPM Network",
        )
        artifacts["cpm_network_mermaid"] = write_cpm_mermaid(
            output_path / "cpm_network.md",
            schedule_rows,
            cpm_rows,
            edge_rows,
            title=f"{title} - CPM Network",
        )
    else:
        warnings.append("schedule_initial/cpm_analysis is empty; CPM visuals were skipped.")

    if resource_load_rows:
        artifacts["resource_load_heatmap"] = write_resource_load_heatmap(
            output_path / "resource_load_heatmap.png",
            resource_load_rows,
            title=f"{title} - Resource Load Rate",
        )
        artifacts["resource_load_bars"] = write_resource_load_bars(
            output_path / "resource_load_bars.png",
            resource_load_rows,
            title=f"{title} - Demand vs Capacity",
        )
    else:
        warnings.append("resource_load_daily is empty; resource load charts were skipped.")

    artifacts["report"] = write_visualization_report(
        output_path / "schedule_visualization_report.md",
        artifacts=artifacts,
        schedule_rows=schedule_rows,
        cpm_rows=cpm_rows,
        resource_load_rows=resource_load_rows,
        warnings=warnings,
        title=title,
    )
    artifacts["manifest"] = write_visualization_manifest(
        output_path / "visualization_manifest.json",
        result=VisualizationResult(output_dir=output_path, artifacts=artifacts, warnings=warnings),
        counts={
            "schedule_initial": len(schedule_rows),
            "cpm_analysis": len(cpm_rows),
            "network_edges": len(edge_rows),
            "resource_load_daily": len(resource_load_rows),
        },
    )
    return VisualizationResult(output_dir=output_path, artifacts=artifacts, warnings=warnings)


def read_exported_schedule_tables(schedule_dir: str | Path) -> dict[str, list[dict[str, Any]]]:
    """Read visualizable tables from the workflow's exported schedule directory."""

    directory = Path(schedule_dir)
    file_by_table = {
        "schedule_initial": "初始施工进度计划.xlsx",
        "cpm_analysis": "关键线路分析表.xlsx",
        "network_edges": "网络计划关系表.xlsx",
        "resource_load_daily": "资源负荷图.xlsx",
    }
    rows: dict[str, list[dict[str, Any]]] = {}
    for table_name, filename in file_by_table.items():
        path = directory / filename
        rows[table_name] = _read_xlsx_rows(path) if path.exists() else []
    if not rows["resource_load_daily"]:
        resource_rows = _read_xlsx_rows(directory / "资源需求表.xlsx")
        if resource_rows:
            rows["resource_load_daily"] = build_resource_load(resource_rows)
    return rows


def build_demo_visualization_rows() -> dict[str, list[dict[str, Any]]]:
    """Build realistic fake schedule rows for visual testing."""

    wbs_rows = [
        _demo_wbs("TASK-001", "1.1", "Preparation", "Site mobilization", 5, ""),
        _demo_wbs("TASK-002", "1.2", "Preparation", "Survey and setting out", 3, "TASK-001"),
        _demo_wbs("TASK-003", "2.1", "Earthwork", "Foundation pit excavation", 12, "TASK-002"),
        _demo_wbs("TASK-004", "2.2", "Earthwork", "Foundation pit support", 10, "TASK-002"),
        _demo_wbs("TASK-005", "3.1", "Basement", "Basement waterproofing", 6, "TASK-003,TASK-004"),
        _demo_wbs("TASK-006", "3.2", "Basement", "Basement structure", 14, "TASK-005"),
        _demo_wbs("TASK-007", "4.1", "Superstructure", "First to fourth floor structure", 18, "TASK-006"),
        _demo_wbs("TASK-008", "4.2", "Superstructure", "Secondary structure", 12, "TASK-006"),
        _demo_wbs("TASK-009", "5.1", "MEP", "MEP rough-in", 14, "TASK-007"),
        _demo_wbs("TASK-010", "5.2", "Finishes", "Interior finishes", 16, "TASK-008,TASK-009"),
        _demo_wbs("TASK-011", "6.1", "Outdoor", "Outdoor utilities and pavement", 10, "TASK-009"),
        _demo_wbs("TASK-012", "7.1", "Acceptance", "Commissioning and acceptance", 5, "TASK-010,TASK-011"),
    ]
    resource_rows = [
        _demo_resource("TASK-003", "labor", "Excavation crew", 36, 30, "2026-W10"),
        _demo_resource("TASK-004", "labor", "Support crew", 28, 32, "2026-W10"),
        _demo_resource("TASK-005", "material", "Waterproofing membrane", 850, 1000, "2026-W12"),
        _demo_resource("TASK-006", "labor", "Rebar crew", 42, 40, "2026-W13"),
        _demo_resource("TASK-006", "equipment", "Tower crane", 1, 1, "2026-W13"),
        _demo_resource("TASK-007", "labor", "Concrete crew", 48, 45, "2026-W15"),
        _demo_resource("TASK-008", "labor", "Masonry crew", 30, 36, "2026-W15"),
        _demo_resource("TASK-009", "labor", "MEP crew", 32, 35, "2026-W18"),
        _demo_resource("TASK-010", "labor", "Finishing crew", 52, 45, "2026-W20"),
        _demo_resource("TASK-011", "equipment", "Excavator", 2, 1, "2026-W20"),
        _demo_resource("TASK-012", "labor", "Commissioning team", 18, 24, "2026-W22"),
    ]
    schedule_rows, cpm_rows, edge_rows = build_initial_schedule(
        wbs_rows,
        resource_rows,
        start_date=date(2026, 3, 1),
    )
    schedule_by_id = {row["task_id"]: row for row in schedule_rows}
    for row in resource_rows:
        schedule = schedule_by_id.get(str(row["task_id"]))
        if not schedule:
            continue
        start = _parse_date(schedule["planned_start"]) or date(2026, 3, 1)
        week = start + timedelta(days=(7 - start.weekday()) % 7)
        row["period"] = f"{week.isocalendar().year}-W{week.isocalendar().week:02d}"
    return {
        "schedule_initial": schedule_rows,
        "cpm_analysis": cpm_rows,
        "network_edges": edge_rows,
        "resource_load_daily": build_resource_load(resource_rows),
    }


def write_gantt_chart(path: Path, rows: list[dict[str, Any]], *, title: str) -> Path:
    """Write a horizontal Gantt chart PNG."""

    tasks = [_schedule_task(row) for row in rows]
    tasks = [task for task in tasks if task["start"] and task["finish"]]
    if not tasks:
        raise ValueError("No schedulable rows with planned_start/planned_finish were found.")
    tasks.sort(key=lambda task: (task["start"], task["finish"], task["task_id"]))

    height = _figure_height(len(tasks), row_height=0.32, minimum=6.5)
    fig, ax = plt.subplots(figsize=(16, height), constrained_layout=True)
    y_positions = range(len(tasks))
    colors = [CRITICAL_COLOR if task["critical"] else NORMAL_COLOR for task in tasks]

    for y, task, color in zip(y_positions, tasks, colors, strict=True):
        start_num = mdates.date2num(task["start"])
        width = max(1, (task["finish"] - task["start"]).days + 1)
        ax.barh(y, width, left=start_num, height=0.62, color=color, edgecolor="white")
        if task["duration"]:
            ax.text(
                start_num + width + 0.5,
                y,
                f"{task['duration']}d",
                va="center",
                fontsize=7,
                color=TEXT_COLOR,
            )

    ax.set_yticks(list(y_positions))
    ax.set_yticklabels([_task_label(task) for task in tasks], fontsize=7)
    ax.invert_yaxis()
    ax.xaxis_date()
    ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=12))
    ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(ax.xaxis.get_major_locator()))
    ax.grid(axis="x", color=GRID_COLOR, linewidth=0.8)
    ax.set_title(title, fontsize=14, pad=12)
    ax.set_xlabel("Calendar")
    _add_critical_legend(ax)
    _save_figure(fig, path)
    return path


def write_cpm_float_chart(
    path: Path,
    schedule_rows: list[dict[str, Any]],
    cpm_rows: list[dict[str, Any]],
    *,
    title: str,
) -> Path:
    """Write a total/free float bar chart PNG."""

    schedule_by_id = {str(row.get("task_id") or ""): row for row in schedule_rows}
    items = []
    for row in cpm_rows:
        task_id = str(row.get("task_id") or "")
        if not task_id:
            continue
        items.append(
            {
                "task_id": task_id,
                "task_name": str(schedule_by_id.get(task_id, {}).get("task_name") or ""),
                "total_float": _as_float(row.get("total_float")),
                "free_float": _as_float(row.get("free_float")),
                "critical": _to_bool(row.get("is_critical")),
            }
        )
    if not items:
        raise ValueError("No CPM rows were found.")
    items.sort(key=lambda item: (item["critical"], -item["total_float"], item["task_id"]))

    height = _figure_height(len(items), row_height=0.28, minimum=6.5)
    fig, ax = plt.subplots(figsize=(14, height), constrained_layout=True)
    y_positions = range(len(items))
    ax.barh(
        list(y_positions),
        [item["total_float"] for item in items],
        color=[CRITICAL_COLOR if item["critical"] else FLOAT_COLOR for item in items],
        height=0.62,
        label="Total float",
    )
    ax.barh(
        list(y_positions),
        [item["free_float"] for item in items],
        color="#9CD49C",
        height=0.28,
        label="Free float",
    )
    ax.set_yticks(list(y_positions))
    ax.set_yticklabels([_task_label(item) for item in items], fontsize=7)
    ax.invert_yaxis()
    ax.grid(axis="x", color=GRID_COLOR, linewidth=0.8)
    ax.set_xlabel("Days")
    ax.set_title(title, fontsize=14, pad=12)
    ax.legend(loc="lower right")
    _save_figure(fig, path)
    return path


def write_cpm_network_chart(
    path: Path,
    schedule_rows: list[dict[str, Any]],
    cpm_rows: list[dict[str, Any]],
    edge_rows: list[dict[str, Any]],
    *,
    title: str,
) -> Path:
    """Write a CPM precedence network PNG."""

    graph, task_meta, critical_edges = _build_network(schedule_rows, cpm_rows, edge_rows)
    if not graph.nodes:
        raise ValueError("No CPM network nodes were found.")

    pos = _cpm_layout(graph, task_meta)
    node_colors = [
        CRITICAL_COLOR if task_meta[node]["critical"] else NORMAL_COLOR for node in graph.nodes
    ]
    node_sizes = [1050 if task_meta[node]["critical"] else 900 for node in graph.nodes]
    edge_colors = [
        CRITICAL_COLOR if (source, target) in critical_edges else "#7A869A"
        for source, target in graph.edges
    ]
    edge_widths = [2.0 if (source, target) in critical_edges else 1.0 for source, target in graph.edges]
    labels = {node: _network_label(node, task_meta[node]["task_name"]) for node in graph.nodes}

    width = min(max(12, len({task_meta[node]["es"] for node in graph.nodes}) * 1.15), 24)
    height = min(max(8, len(graph.nodes) * 0.18 + 4), 22)
    fig, ax = plt.subplots(figsize=(width, height), constrained_layout=True)
    nx.draw_networkx_edges(
        graph,
        pos,
        ax=ax,
        edge_color=edge_colors,
        width=edge_widths,
        arrows=True,
        arrowsize=12,
        connectionstyle="arc3,rad=0.05",
    )
    nx.draw_networkx_nodes(
        graph,
        pos,
        ax=ax,
        node_color=node_colors,
        node_size=node_sizes,
        edgecolors="white",
        linewidths=1.3,
    )
    nx.draw_networkx_labels(
        graph,
        pos,
        labels=labels,
        ax=ax,
        font_size=6.5,
        font_color="white",
    )
    ax.set_title(title, fontsize=14, pad=12)
    ax.axis("off")
    _add_critical_legend(ax)
    _save_figure(fig, path)
    return path


def write_resource_load_heatmap(
    path: Path,
    rows: list[dict[str, Any]],
    *,
    title: str,
) -> Path:
    """Write a resource load-rate heatmap PNG."""

    matrix, resources, periods = _resource_matrix(rows)
    if not resources or not periods:
        raise ValueError("No plottable resource load rows were found.")

    fig_width = min(max(10, len(periods) * 0.55 + 4), 22)
    fig_height = min(max(6, len(resources) * 0.42 + 2.5), 20)
    fig, ax = plt.subplots(figsize=(fig_width, fig_height), constrained_layout=True)
    image = ax.imshow(matrix, aspect="auto", cmap="RdYlGn_r", vmin=0, vmax=max(1.5, matrix.max()))
    ax.set_xticks(range(len(periods)))
    ax.set_xticklabels(periods, rotation=45, ha="right", fontsize=8)
    ax.set_yticks(range(len(resources)))
    ax.set_yticklabels(resources, fontsize=8)
    ax.set_title(title, fontsize=14, pad=12)
    ax.set_xlabel("Period")
    ax.set_ylabel("Resource")
    colorbar = fig.colorbar(image, ax=ax)
    colorbar.set_label("Load rate")
    if len(resources) * len(periods) <= 240:
        for y, row in enumerate(matrix):
            for x, value in enumerate(row):
                if value > 0:
                    ax.text(
                        x,
                        y,
                        f"{value:.2f}",
                        ha="center",
                        va="center",
                        fontsize=6,
                        color="white" if value >= 1 else TEXT_COLOR,
                    )
    _save_figure(fig, path)
    return path


def write_resource_load_bars(
    path: Path,
    rows: list[dict[str, Any]],
    *,
    title: str,
) -> Path:
    """Write a top resource demand-vs-capacity bar chart PNG."""

    items = []
    for row in rows:
        demand = _as_float(row.get("demand"))
        capacity = _as_float(row.get("capacity"))
        if not row.get("resource_name") or not row.get("date_or_period"):
            continue
        items.append(
            {
                "label": _truncate(
                    f"{row.get('resource_name')} | {row.get('date_or_period')}",
                    42,
                ),
                "demand": demand,
                "capacity": capacity,
                "load_rate": demand / capacity if capacity else math.inf,
                "conflict": _to_bool(row.get("conflict_flag")) or (capacity and demand > capacity),
            }
        )
    if not items:
        raise ValueError("No plottable resource load rows were found.")
    items.sort(key=lambda item: (-item["conflict"], -item["load_rate"], item["label"]))
    items = items[:25]

    height = _figure_height(len(items), row_height=0.36, minimum=6)
    fig, ax = plt.subplots(figsize=(14, height), constrained_layout=True)
    y_positions = list(range(len(items)))
    ax.barh(
        y_positions,
        [item["capacity"] for item in items],
        color=CAPACITY_COLOR,
        alpha=0.55,
        height=0.7,
        label="Capacity",
    )
    ax.barh(
        y_positions,
        [item["demand"] for item in items],
        color=[CRITICAL_COLOR if item["conflict"] else NORMAL_COLOR for item in items],
        height=0.42,
        label="Demand",
    )
    ax.set_yticks(y_positions)
    ax.set_yticklabels([item["label"] for item in items], fontsize=8)
    ax.invert_yaxis()
    ax.grid(axis="x", color=GRID_COLOR, linewidth=0.8)
    ax.set_xlabel("Quantity")
    ax.set_title(title, fontsize=14, pad=12)
    ax.legend(loc="lower right")
    _save_figure(fig, path)
    return path


def write_cpm_mermaid(
    path: Path,
    schedule_rows: list[dict[str, Any]],
    cpm_rows: list[dict[str, Any]],
    edge_rows: list[dict[str, Any]],
    *,
    title: str,
) -> Path:
    """Write a Mermaid CPM diagram for Markdown-friendly previews."""

    cpm_by_id = {str(row.get("task_id") or ""): row for row in cpm_rows}
    node_ids = {
        str(row.get("task_id") or ""): str(row.get("task_id") or "").replace("-", "_")
        for row in schedule_rows
        if row.get("task_id")
    }
    lines = [f"# {title}", "", "```mermaid", "flowchart LR"]
    for row in schedule_rows:
        task_id = str(row.get("task_id") or "")
        if not task_id:
            continue
        cpm = cpm_by_id.get(task_id, {})
        label = _mermaid_label(
            f"{task_id}<br/>{row.get('task_name') or ''}<br/>TF={cpm.get('total_float', '')}"
        )
        lines.append(f'  {node_ids[task_id]}["{label}"]')
    for edge in edge_rows:
        source = str(edge.get("from_task_id") or "")
        target = str(edge.get("to_task_id") or "")
        if source not in node_ids or target not in node_ids:
            continue
        connector = "==>" if _to_bool(edge.get("is_critical_edge")) else "-->"
        lines.append(f"  {node_ids[source]} {connector} {node_ids[target]}")
    critical_nodes = [
        node_ids[str(row.get("task_id") or "")]
        for row in cpm_rows
        if _to_bool(row.get("is_critical")) and str(row.get("task_id") or "") in node_ids
    ]
    if critical_nodes:
        lines.append(f"  class {' '.join(critical_nodes)} critical")
        lines.append("  classDef critical fill:#ffd6d3,stroke:#d94841,stroke-width:2px")
    lines.extend(["```", ""])
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def write_visualization_report(
    path: Path,
    *,
    artifacts: dict[str, Path],
    schedule_rows: list[dict[str, Any]],
    cpm_rows: list[dict[str, Any]],
    resource_load_rows: list[dict[str, Any]],
    warnings: list[str],
    title: str,
) -> Path:
    """Write a compact Markdown index for generated visualizations."""

    critical_count = sum(1 for row in cpm_rows if _to_bool(row.get("is_critical")))
    conflict_count = sum(1 for row in resource_load_rows if _to_bool(row.get("conflict_flag")))
    lines = [
        f"# {title}",
        "",
        "## Summary",
        "",
        f"- Schedule tasks: `{len(schedule_rows)}`",
        f"- Critical tasks: `{critical_count}`",
        f"- Resource load rows: `{len(resource_load_rows)}`",
        f"- Resource conflicts: `{conflict_count}`",
        "",
        "## Artifacts",
        "",
    ]
    for name, artifact_path in artifacts.items():
        if name in {"report", "manifest"}:
            continue
        lines.append(f"- `{name}`: `{artifact_path.name}`")
    if warnings:
        lines.extend(["", "## Warnings", ""])
        lines.extend(f"- {warning}" for warning in warnings)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_visualization_manifest(
    path: Path,
    *,
    result: VisualizationResult,
    counts: dict[str, int],
) -> Path:
    """Write machine-readable visualization metadata."""

    payload = result.to_dict()
    payload["counts"] = counts
    payload["created_at"] = datetime.now().isoformat(timespec="seconds")
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def _schedule_task(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "task_id": str(row.get("task_id") or ""),
        "task_name": str(row.get("task_name") or ""),
        "phase": str(row.get("phase") or ""),
        "start": _parse_date(row.get("planned_start")),
        "finish": _parse_date(row.get("planned_finish")),
        "duration": int(_as_float(row.get("duration_days"))),
        "critical": _to_bool(row.get("is_critical")),
    }


def _demo_wbs(
    task_id: str,
    wbs_code: str,
    phase: str,
    task_name: str,
    duration_days: int,
    predecessor_ids: str,
) -> dict[str, Any]:
    return {
        "task_id": task_id,
        "wbs_code": wbs_code,
        "phase": phase,
        "section": "",
        "floor_or_area": "",
        "task_name": task_name,
        "work_package": phase,
        "quantity": "",
        "unit": "",
        "duration_days": duration_days,
        "predecessor_ids": predecessor_ids,
        "relation_type": "FS",
        "lag_days": 0,
        "source": "demo data",
        "confidence": "1.00",
        "note": "Fake row for visualization smoke testing.",
        "owner_agent": "visualization_demo",
    }


def _demo_resource(
    task_id: str,
    resource_type: str,
    resource_name: str,
    demand: float,
    capacity: float,
    period: str,
) -> dict[str, Any]:
    return {
        "task_id": task_id,
        "resource_type": resource_type,
        "resource_name": resource_name,
        "demand": demand,
        "unit": "unit",
        "capacity": capacity,
        "period": period,
        "conflict_flag": demand > capacity,
        "source": "demo data",
        "confidence": "1.00",
        "note": "Fake row for visualization smoke testing.",
        "owner_agent": "visualization_demo",
    }


def _build_network(
    schedule_rows: list[dict[str, Any]],
    cpm_rows: list[dict[str, Any]],
    edge_rows: list[dict[str, Any]],
) -> tuple[nx.DiGraph, dict[str, dict[str, Any]], set[tuple[str, str]]]:
    cpm_by_id = {str(row.get("task_id") or ""): row for row in cpm_rows}
    graph = nx.DiGraph()
    task_meta: dict[str, dict[str, Any]] = {}
    for row in schedule_rows:
        task_id = str(row.get("task_id") or "")
        if not task_id:
            continue
        cpm = cpm_by_id.get(task_id, {})
        meta = {
            "task_name": str(row.get("task_name") or ""),
            "es": int(_as_float(cpm.get("es"))),
            "ef": int(_as_float(cpm.get("ef"))),
            "critical": _to_bool(cpm.get("is_critical") or row.get("is_critical")),
        }
        graph.add_node(task_id)
        task_meta[task_id] = meta

    critical_edges: set[tuple[str, str]] = set()
    for edge in edge_rows:
        source = str(edge.get("from_task_id") or "")
        target = str(edge.get("to_task_id") or "")
        if source in task_meta and target in task_meta:
            graph.add_edge(source, target)
            if _to_bool(edge.get("is_critical_edge")):
                critical_edges.add((source, target))
    return graph, task_meta, critical_edges


def _cpm_layout(graph: nx.DiGraph, task_meta: dict[str, dict[str, Any]]) -> dict[str, tuple[float, float]]:
    groups: dict[int, list[str]] = {}
    for node in graph.nodes:
        groups.setdefault(task_meta[node]["es"], []).append(node)
    pos: dict[str, tuple[float, float]] = {}
    for x_index, es in enumerate(sorted(groups)):
        nodes = sorted(groups[es], key=lambda node: (task_meta[node]["ef"], node))
        center = (len(nodes) - 1) / 2
        for index, node in enumerate(nodes):
            pos[node] = (x_index, center - index)
    return pos


def _resource_matrix(rows: list[dict[str, Any]]):
    periods = sorted({str(row.get("date_or_period") or "") for row in rows if row.get("date_or_period")})
    resources = sorted(
        {str(row.get("resource_name") or "") for row in rows if row.get("resource_name")}
    )
    max_load_by_resource = {
        resource: max(
            (
                _as_float(row.get("load_rate"))
                for row in rows
                if str(row.get("resource_name") or "") == resource
            ),
            default=0,
        )
        for resource in resources
    }
    resources = sorted(resources, key=lambda resource: (-max_load_by_resource[resource], resource))[:30]
    period_index = {period: index for index, period in enumerate(periods)}
    resource_index = {resource: index for index, resource in enumerate(resources)}

    import numpy as np

    matrix = np.zeros((len(resources), len(periods)))
    for row in rows:
        resource = str(row.get("resource_name") or "")
        period = str(row.get("date_or_period") or "")
        if resource not in resource_index or period not in period_index:
            continue
        matrix[resource_index[resource], period_index[period]] = max(
            matrix[resource_index[resource], period_index[period]],
            _as_float(row.get("load_rate")),
        )
    return matrix, [_truncate(resource, 36) for resource in resources], periods


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    output: list[dict[str, Any]] = []
    for values in rows[1:]:
        if not values or all(value is None for value in values):
            continue
        output.append(
            {
                headers[index]: value
                for index, value in enumerate(values[: len(headers)])
                if headers[index]
            }
        )
    return output


def _as_float(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是"}


def _truncate(value: str, length: int) -> str:
    return value if len(value) <= length else value[: max(length - 3, 0)] + "..."


def _task_label(task: dict[str, Any]) -> str:
    return _truncate(f"{task['task_id']} {task['task_name']}", 48)


def _network_label(task_id: str, task_name: str) -> str:
    name = _truncate(task_name, 16)
    return f"{task_id}\n{name}" if name else task_id


def _mermaid_label(value: str) -> str:
    return value.replace('"', "'").replace("[", "(").replace("]", ")")


def _figure_height(count: int, *, row_height: float, minimum: float) -> float:
    return min(max(minimum, count * row_height + 2.5), 28)


def _add_critical_legend(ax: plt.Axes) -> None:
    handles = [
        plt.Line2D([0], [0], marker="s", color="w", label="Critical", markerfacecolor=CRITICAL_COLOR, markersize=9),
        plt.Line2D([0], [0], marker="s", color="w", label="Non-critical", markerfacecolor=NORMAL_COLOR, markersize=9),
    ]
    ax.legend(handles=handles, loc="lower right")


def _configure_matplotlib() -> None:
    plt.rcParams["axes.unicode_minus"] = False
    preferred_fonts = (
        "Microsoft YaHei",
        "SimHei",
        "Noto Sans CJK SC",
        "Source Han Sans SC",
        "Arial Unicode MS",
    )
    available = {font.name for font in font_manager.fontManager.ttflist}
    for font in preferred_fonts:
        if font in available:
            plt.rcParams["font.sans-serif"] = [font, "DejaVu Sans"]
            return
    plt.rcParams["font.sans-serif"] = ["DejaVu Sans"]


def _save_figure(fig: plt.Figure, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)
