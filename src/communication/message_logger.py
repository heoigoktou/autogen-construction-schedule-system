"""Message logging utilities."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from blackboard.excel_store import ExcelBlackboardStore
from blackboard.sheet_schema import get_sheet_spec
from communication.message_schema import AgentMessage


class MessageLogger:
    """Write Agent communication records to the blackboard and export logs."""

    def __init__(self, store: ExcelBlackboardStore, communication_log_path: str | Path) -> None:
        self.store = store
        self.communication_log_path = Path(communication_log_path)

    def log(self, message: AgentMessage) -> None:
        """Append one message to `agent_message_log`."""

        self.store.append_row("agent_message_log", message.to_log_row())

    def export_communication_log(self) -> Path:
        """Export `agent_message_log` to the configured communication log workbook."""

        self.communication_log_path.parent.mkdir(parents=True, exist_ok=True)
        rows = self.store.read_rows("agent_message_log")
        spec = get_sheet_spec("agent_message_log")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "agent_message_log"
        sheet.append(list(spec.headers))
        for row in rows:
            sheet.append([row.get(header) for header in spec.headers])
        workbook.save(self.communication_log_path)
        return self.communication_log_path

    def count_messages(self) -> int:
        """Return current number of logged messages."""

        return len(self.store.read_rows("agent_message_log"))

    def load_exported_rows(self) -> list[dict[str, object]]:
        """Read exported communication log rows, mainly for smoke checks."""

        if not self.communication_log_path.exists():
            return []
        workbook = load_workbook(self.communication_log_path, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        result = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if values and any(value is not None for value in values):
                result.append({headers[index]: value for index, value in enumerate(values)})
        return result
