"""Excel public blackboard store.

All formal data exchange goes through this workbook abstraction. No SQL or
external database is used.
"""

from __future__ import annotations

import os
import threading
import uuid
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from blackboard.sheet_schema import SHEET_SPECS, all_sheet_names, get_sheet_spec
from blackboard.validators import validate_headers, validate_row


class ExcelBlackboardStore:
    """Read and write the shared Excel public blackboard."""

    def __init__(self, workbook_path: str | Path) -> None:
        self.workbook_path = Path(workbook_path)
        self._lock = threading.RLock()

    def initialize(self, overwrite: bool = False) -> Path:
        """Create the public blackboard workbook with all required sheets."""

        with self._lock:
            if self.workbook_path.exists() and not overwrite:
                self.ensure_schema()
                return self.workbook_path

            self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            for sheet_name in all_sheet_names():
                spec = get_sheet_spec(sheet_name)
                sheet = workbook.create_sheet(sheet_name)
                sheet.append(list(spec.headers))
                self._style_header(sheet)
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = f"A1:{get_column_letter(len(spec.headers))}1"
                self._set_reasonable_widths(sheet, spec.headers)

            self._save_workbook(workbook)
            return self.workbook_path

    def ensure_schema(self) -> None:
        """Ensure every required sheet exists and contains required headers."""

        with self._lock:
            if not self.workbook_path.exists():
                self.initialize(overwrite=False)
                return

            workbook = load_workbook(self.workbook_path)
            changed = False
            for sheet_name, spec in SHEET_SPECS.items():
                if sheet_name not in workbook.sheetnames:
                    sheet = workbook.create_sheet(sheet_name)
                    sheet.append(list(spec.headers))
                    self._style_header(sheet)
                    self._set_reasonable_widths(sheet, spec.headers)
                    changed = True
                    continue

                sheet = workbook[sheet_name]
                headers = self._read_headers_from_sheet(sheet)
                if not headers:
                    sheet.append(list(spec.headers))
                    self._style_header(sheet)
                    self._set_reasonable_widths(sheet, spec.headers)
                    changed = True
                    headers = list(spec.headers)
                missing_headers = [header for header in spec.headers if header not in headers]
                if missing_headers:
                    for header in missing_headers:
                        sheet.cell(row=1, column=sheet.max_column + 1, value=header)
                    headers = self._read_headers_from_sheet(sheet)
                    self._style_header(sheet)
                    self._set_reasonable_widths(sheet, tuple(headers))
                    sheet.freeze_panes = "A2"
                    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
                    changed = True
                validate_headers(sheet_name, headers)

            if changed:
                self._save_workbook(workbook)

    def read_headers(self, sheet_name: str) -> list[str]:
        """Read the first row from a worksheet as headers."""

        with self._lock:
            workbook = self._load()
            sheet = workbook[sheet_name]
            headers = self._read_headers_from_sheet(sheet)
            validate_headers(sheet_name, headers)
            return headers

    def read_rows(self, sheet_name: str) -> list[dict[str, Any]]:
        """Read all data rows from a worksheet as dictionaries."""

        with self._lock:
            workbook = self._load()
            sheet = workbook[sheet_name]
            headers = self._read_headers_from_sheet(sheet)
            validate_headers(sheet_name, headers)
            rows: list[dict[str, Any]] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                if values is None or all(value is None for value in values):
                    continue
                rows.append(
                    {headers[index]: value for index, value in enumerate(values[: len(headers)])}
                )
            return rows

    def append_row(self, sheet_name: str, row: Mapping[str, Any]) -> None:
        """Append one validated row to a worksheet."""

        self.append_rows(sheet_name, [row])

    def append_rows(self, sheet_name: str, rows: Iterable[Mapping[str, Any]]) -> None:
        """Append multiple validated rows to a worksheet."""

        with self._lock:
            self.ensure_schema()
            workbook = self._load()
            sheet = workbook[sheet_name]
            headers = self._read_headers_from_sheet(sheet)
            validate_headers(sheet_name, headers)

            for row in rows:
                validate_row(sheet_name, row)
                sheet.append([row.get(header) for header in headers])

            self._style_header(sheet)
            self._save_workbook(workbook)

    def replace_rows(self, sheet_name: str, rows: Iterable[Mapping[str, Any]]) -> None:
        """Replace data rows in one worksheet while preserving the agreed schema."""

        with self._lock:
            self.ensure_schema()
            workbook = self._load()
            sheet = workbook[sheet_name]
            headers = self._read_headers_from_sheet(sheet)
            validate_headers(sheet_name, headers)

            if sheet.max_row > 1:
                sheet.delete_rows(2, sheet.max_row - 1)

            for row in rows:
                validate_row(sheet_name, row)
                sheet.append([row.get(header) for header in headers])

            self._style_header(sheet)
            self._save_workbook(workbook)

    def replace_sheets_rows(
        self,
        sheet_rows: Mapping[str, Iterable[Mapping[str, Any]]],
    ) -> None:
        """Replace rows in multiple worksheets and save the workbook once."""

        with self._lock:
            self.ensure_schema()
            workbook = self._load()
            for sheet_name, rows in sheet_rows.items():
                sheet = workbook[sheet_name]
                headers = self._read_headers_from_sheet(sheet)
                validate_headers(sheet_name, headers)

                if sheet.max_row > 1:
                    sheet.delete_rows(2, sheet.max_row - 1)

                for row in rows:
                    validate_row(sheet_name, row)
                    sheet.append([row.get(header) for header in headers])

                self._style_header(sheet)
            self._save_workbook(workbook)

    def copy_parameter_template(self, template_path: str | Path) -> int:
        """Import the parameter checklist template into the demo blackboard.

        The workbook template uses Chinese field names. This method maps the
        available columns to the normalized `parameter_checklist` sheet.
        """

        template = Path(template_path)
        if not template.exists():
            return 0

        source_wb = load_workbook(template, data_only=True)
        if "参数检查清单" not in source_wb.sheetnames:
            return 0

        source_sheet = source_wb["参数检查清单"]
        headers = [cell.value for cell in next(source_sheet.iter_rows(min_row=1, max_row=1))]
        header_index = {name: idx for idx, name in enumerate(headers) if name}

        imported: list[dict[str, Any]] = []
        for values in source_sheet.iter_rows(min_row=2, values_only=True):
            if not values or not values[0]:
                continue
            imported.append(
                {
                    "parameter_id": values[header_index.get("参数编号", 0)],
                    "category": values[header_index.get("参数类别", 1)],
                    "name": values[header_index.get("参数名称", 2)],
                    "required": (
                        values[header_index.get("是否必需", 5)]
                        if "是否必需" in header_index
                        else "是"
                    ),
                    "source": (
                        values[header_index.get("优先来源", 7)]
                        if "优先来源" in header_index
                        else ""
                    ),
                    "status": (
                        values[header_index.get("检查状态", 12)]
                        if "检查状态" in header_index
                        else "待检查"
                    ),
                    "owner_agent": (
                        values[header_index.get("使用Agent", 4)]
                        if "使用Agent" in header_index
                        else "data_parser_agent"
                    ),
                    "note": values[header_index.get("备注", 13)] if "备注" in header_index else "",
                    "created_by": "system_import",
                    "created_at": "",
                }
            )

        if imported:
            self.append_rows("parameter_checklist", imported)
        return len(imported)

    def _load(self):
        self.ensure_schema()
        return load_workbook(self.workbook_path)

    @staticmethod
    def _read_headers_from_sheet(sheet) -> list[str]:
        if sheet.max_row < 1:
            return []
        return [cell.value for cell in sheet[1] if cell.value]

    @staticmethod
    def _style_header(sheet) -> None:
        fill = PatternFill("solid", fgColor="1F4E79")
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    @staticmethod
    def _set_reasonable_widths(sheet, headers: tuple[str, ...]) -> None:
        for index, header in enumerate(headers, start=1):
            width = min(max(len(header) + 4, 14), 28)
            sheet.column_dimensions[get_column_letter(index)].width = width

    def _save_workbook(self, workbook) -> None:
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = self.workbook_path.with_name(
            f"{self.workbook_path.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp"
        )
        workbook.save(tmp_path)
        os.replace(tmp_path, self.workbook_path)

    def export_sheet_copy(self, output_path: str | Path, sheet_name: str) -> Path:
        """Export one worksheet into a small standalone workbook."""

        rows = self.read_rows(sheet_name)
        spec = get_sheet_spec(sheet_name)
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(list(spec.headers))
        for row in rows:
            sheet.append([row.get(header) for header in spec.headers])
        self._style_header(sheet)
        workbook.save(output)
        return output
